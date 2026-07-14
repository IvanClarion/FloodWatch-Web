import os
import json
from pathlib import Path

class LantawExtractDataFromFiles:
    """
    Extracts structured text and tabular data from document-type files
    for the Lantaw AI pipeline.

    Supported formats:
      - .pdf   → Text extraction via PyPDF2 / pdfplumber
      - .docx  → Text extraction via python-docx
      - .xlsx  → Tabular extraction via openpyxl
      - .csv   → Tabular extraction via csv module
      - .txt   → Direct text read
    """

    SUPPORTED_EXTENSIONS = [".pdf", ".docx", ".xlsx", ".csv", ".txt"]

    @staticmethod
    def extract(file_path: str) -> dict:
        """
        Main extraction dispatcher. Reads the file and returns its contents
        in a structured dictionary.

        Returns:
            dict with keys: 'type' (str), 'content' (str|list), 'rows' (int|None), 'error' (str|None)
        """
        ext = Path(file_path).suffix.lower()

        if ext not in LantawExtractDataFromFiles.SUPPORTED_EXTENSIONS:
            return {
                "type": "unsupported",
                "content": None,
                "rows": None,
                "error": f"Unsupported document format: {ext}"
            }

        try:
            if ext == ".txt":
                return LantawExtractDataFromFiles._extract_txt(file_path)
            elif ext == ".csv":
                return LantawExtractDataFromFiles._extract_csv(file_path)
            elif ext == ".pdf":
                return LantawExtractDataFromFiles._extract_pdf(file_path)
            elif ext == ".docx":
                return LantawExtractDataFromFiles._extract_docx(file_path)
            elif ext == ".xlsx":
                return LantawExtractDataFromFiles._extract_xlsx(file_path)
        except Exception as e:
            return {
                "type": ext.replace(".", ""),
                "content": None,
                "rows": None,
                "error": f"Error extracting {ext} file: {str(e)}"
            }

    # ── Plain Text ──────────────────────────────────────────────
    @staticmethod
    def _extract_txt(file_path: str) -> dict:
        """Reads a plain text file and returns its full content."""
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
        return {
            "type": "txt",
            "content": content,
            "rows": content.count("\n") + 1,
            "error": None
        }

    # ── CSV ─────────────────────────────────────────────────────
    @staticmethod
    def _extract_csv(file_path: str) -> dict:
        """Reads a CSV file and returns it as a list of row dictionaries."""
        import csv
        rows = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
        return {
            "type": "csv",
            "content": rows,
            "rows": len(rows),
            "error": None
        }

    # ── PDF ─────────────────────────────────────────────────────
    @staticmethod
    def _extract_pdf(file_path: str) -> dict:
        """
        Extracts text from a PDF file using pdfplumber (preferred) 
        or falls back to PyPDF2.
        """
        text = ""

        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except ImportError:
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(file_path)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            except ImportError:
                return {
                    "type": "pdf",
                    "content": None,
                    "rows": None,
                    "error": "No PDF library available. Install pdfplumber or PyPDF2."
                }

        return {
            "type": "pdf",
            "content": text.strip(),
            "rows": text.count("\n") + 1 if text.strip() else 0,
            "error": None
        }

    # ── DOCX ────────────────────────────────────────────────────
    @staticmethod
    def _extract_docx(file_path: str) -> dict:
        """Extracts text from a DOCX file paragraph by paragraph."""
        try:
            from docx import Document
        except ImportError:
            return {
                "type": "docx",
                "content": None,
                "rows": None,
                "error": "python-docx is not installed. Run: pip install python-docx"
            }

        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        content = "\n".join(paragraphs)

        return {
            "type": "docx",
            "content": content,
            "rows": len(paragraphs),
            "error": None
        }

    # ── XLSX ────────────────────────────────────────────────────
    @staticmethod
    def _extract_xlsx(file_path: str) -> dict:
        """
        Extracts data from the first sheet of an XLSX file.
        Returns rows as a list of dictionaries using the first row as headers.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "type": "xlsx",
                "content": None,
                "rows": None,
                "error": "openpyxl is not installed. Run: pip install openpyxl"
            }

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        rows_data = []
        headers = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell else f"col_{j}" for j, cell in enumerate(row)]
            else:
                row_dict = {}
                for j, cell in enumerate(row):
                    key = headers[j] if j < len(headers) else f"col_{j}"
                    row_dict[key] = cell
                rows_data.append(row_dict)

        wb.close()

        return {
            "type": "xlsx",
            "content": rows_data,
            "rows": len(rows_data),
            "error": None
        }

    @staticmethod
    def content_to_context_string(extracted: dict) -> str:
        """
        Converts the extracted dictionary into a clean string that can be
        injected into the AI prompt as context data.
        """
        if extracted.get("error"):
            return f"[File extraction error: {extracted['error']}]"

        content = extracted.get("content")
        file_type = extracted.get("type", "unknown")

        if isinstance(content, list):
            # Tabular data (CSV / XLSX): convert to readable text
            if not content:
                return "[Empty file — no data rows found]"
            preview = content[:50]  # Limit to 50 rows for context
            return (
                f"[Extracted {file_type.upper()} data — {extracted.get('rows', len(content))} rows]\n"
                + json.dumps(preview, indent=2, default=str)
            )
        elif isinstance(content, str):
            # Text data (TXT / PDF / DOCX): truncate if very large
            max_chars = 8000
            truncated = content[:max_chars]
            suffix = "\n...[truncated]" if len(content) > max_chars else ""
            return (
                f"[Extracted {file_type.upper()} text — {extracted.get('rows', 0)} lines]\n"
                + truncated + suffix
            )
        else:
            return "[No content extracted]"
